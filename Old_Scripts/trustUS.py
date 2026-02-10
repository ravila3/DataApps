import threading
import tkinter as tk
# from multiprocessing import freeze_support
from tkinter import filedialog as fd
from tkinter import ttk
import openpyxl
import requests
import time
import os
import sv_ttk
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Replace the required_category to your required category
required_category = 'shopping_fashion'
file_name = f'company-{required_category}.xlsx'

# Check if the file exists
if not os.path.exists(file_name):
    # Create a new workbook and worksheet if the file doesn't exist
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet1'
    # Save the new file
    workbook.save(file_name)
    print(f"{file_name} created successfully.")
else:
    # Open the existing file
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    print(f"{file_name} loaded successfully.")

# Open the Excel file
# workbook = openpyxl.load_workbook(
#     'company-{required_category}.xlsx')
# sheet = workbook.active

def scrape_info(soup, start_row):
    global info_text
    driver_detail = webdriver.Edge()
    # workbook.save(f'Company-{required_category}_update.xlsx')

    business_links = soup.find_all('a', {'name': 'business-unit-card'})

    # Extract the href values
    hrefs = [link['href'] for link in business_links if 'href' in link.attrs]
    print(f"hrefs = {hrefs}") #debug

    # Process each href
    for link in business_links:
        sheet[f'A{start_row}'].value = start_row-1
        company_name=link.find('p', class_='CDS_Typography_heading-xs__bedfe1').text.strip() # xs__jSwUz
        print(f"company name = {company_name}")
        sheet[f'B{start_row}'].value = company_name
        try:
            company_score=link.find('div', class_='styles_rating__AWUbj').text.strip() # styles_rating__pY5Pk
            print(f"company_score = {company_score}")
            sheet[f'D{start_row}'].value = company_score
        except:
            sheet[f'D{start_row}'].value = "TrustScore 0|0 reviews"

        try:
            company_location=link.find('span', class_='styles_location__wea8G').text.strip() #styles_location__ILZb0
            sheet[f'E{start_row}'].value = company_location
            print(f"company location = {company_location}") # debug
        except:
            sheet[f'E{start_row}'].value = "Australia"
            
        href=link['href']
        company_domain=href.replace("/review/", "")
        sheet[f'C{start_row}'].value = company_domain
        business_url = f"https://au.trustpilot.com{href}"
        print(f"Processing {business_url}...")
        time.sleep(0.2)
        # Send a GET request to the individual business page
        response=requests.request("GET", business_url)
        # response_sub = response.text
        # print(BeautifulSoup(response_sub, "html.parser"))
        
        # Check if the request was successful
        if response.status_code == 200:
            # Further processing can be done here with response_sub.content
            print(f"Successfully retrieved data from {business_url}")
            # Set up Selenium WebDriver
            
            driver_detail.get(business_url)
            
            page_source = driver_detail.page_source
            # Let the page load completely
            soup_sub = BeautifulSoup(page_source, "html.parser")
            # driver.quit()  # Close the browser after use
            # print(soup_sub.prettify())

            # exit() #debug
            try:
                categories = []
                for link_tag in soup_sub.find_all('a', class_=lambda class_name: class_name and class_name.endswith('styles_categoriesLink__Rz2T0')):
                # for link_tag in soup_sub.find_all('a', class_='link_internal__Eam_b typography_appearance-action__u_Du4 link_link__jBdLV link_underlined__eziE0 styles_categoriesLink__Rz2T0'):
                    # print(f"link_tag = {link_tag}") #debug
                    categories.append(link_tag.text.strip())
                print(f"categories = {categories}")
                
                # company_activity_card = soup_sub.find('div', class_='styles_companyActivityDesktop__or977')
                # list_items = company_activity_card.find_all('div', class_='styles_listItem__7beWu')
                # company_activiti_string= "".join([list.get_text().strip() + " / " for list in list_items])
                # print(f"company_activity_card = {company_activity_card}, list_items = {list_items}, company_activiti_string = {company_activiti_string}") #debug
            except:
                # company_activiti_string= ""
                categories = []
            
            # exit() #debug
            categories_string = ", ".join(categories)

            sheet[f'F{start_row}'].value = categories_string
            start_row += 1

        else:
            print(f"Failed to retrieve data from {business_url}")

        print("\n")

def process_request():
    driver = webdriver.Chrome()  # You may need to adjust the path to your Chrome driver
    # Launch Chrome browser
    url = f"https://au.trustpilot.com/categories/{required_category}"
    driver.get(url)
    agree_buttons = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CLASS_NAME, "onetrust-close-btn-handler"))
    )
    By.ID
    agree_buttons.click()
    try:
        # Try to find the element with data-pagination-button-last-link='true'
        page = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@data-pagination-button-last-link='true']"))
        ).text
    except:
        try:
            page = WebDriverWait(driver, 1).until(
                EC.element_to_be_clickable((By.XPATH, "//a[@data-pagination-button-4-link='true']"))
            ).text
        except:
            try:
                page = WebDriverWait(driver, 1).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@data-pagination-button-3-link='true']"))
                ).text
            except:
                try:
                    page = WebDriverWait(driver, 1).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@data-pagination-button-2-link='true']"))
                    ).text
                except:
                    page = WebDriverWait(driver, 1).until(
                        EC.element_to_be_clickable((By.XPATH, "//a[@data-pagination-button-1-link='true']"))
                    ).text
    print(f"total pages = {int(page)}")
    
    page = 1 #int(page)
    # Find all pagination buttons
    for idx in range(0, page):
        print(f"Processing page {idx+1}")
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html.parser")
        scrape_info(soup, start_row=idx * 20 + 2)
        # Check if it's not the last page
        if idx < page:
            button = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH,"//a[@data-pagination-button-next-link='true']")))
            driver.execute_script("arguments[0].click();", button)
            # Wait for the new content to load
            WebDriverWait(driver, 20)

            # Optionally, add a short delay to ensure the content is fully loaded
            time.sleep(5)
        # workbook.save(f'Company-{required_category}-{idx}.xlsx')
    driver.quit()  # Close the browser after scraping

def main():

    process_request()
    # Save the modified workbook

    try:
        workbook.save(f'Company-{required_category}.xlsx')
        # Close the workbook
        workbook.close()
    except FileNotFoundError:
        print("file cannot found!")
        return


if __name__ == '__main__':
    main()

